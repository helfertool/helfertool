from django.template import defaultfilters as filters
from django.utils.translation import ugettext as _
import openpyxl

from .xlsxutils import CellIterator, cleanName, escape


def xlsx_helpers_in_job(buffer, event, jobs, date):
    """ Exports the helpers for given jobs of an event as excel spreadsheet.

    Parameter:
        buffer: a writeable bytes buffer (e.g. io.BytesIO or a file)
        event:  the exported event
        jobs:   a list of all exported jobs
    """
    # create xlsx
    workbook = openpyxl.Workbook()

    # remove the default created worksheet
    workbook.remove(workbook.active)

    # duplicated worksheet names are not allowed
    used_names = []

    # styles
    bold = openpyxl.styles.Font(bold=True)
    multiple_shifts = openpyxl.styles.PatternFill(
        start_color='FFFFFF99',
        end_color='FFfFFF99',
        fill_type='solid')

    # export jobs
    for job in jobs:
        job_name = cleanName(job.name)[:20]  # worksheet name must be <= 31 chars

        job_name_use = job_name
        counter = 2
        while job_name_use in used_names:
            job_name_use = "{}{}".format(job_name, counter)
            counter += 1

        used_names.append(job_name_use)

        worksheet = workbook.create_sheet(title=job_name_use)
        iterator = CellIterator(worksheet)

        worksheet.column_dimensions[iterator.column].width = 30
        iterator.write(_("First name"), bold).right()

        worksheet.column_dimensions[iterator.column].width = 30
        iterator.write(_("Surname"), bold).right()

        worksheet.column_dimensions[iterator.column].width = 30
        iterator.write(_("E-Mail"), bold).right()

        if event.ask_phone:
            worksheet.column_dimensions[iterator.column].width = 20
            iterator.write(_("Mobile phone"), bold).right()

        if event.ask_shirt:
            worksheet.column_dimensions[iterator.column].width = 10
            iterator.write(_("T-shirt"), bold).right()

        if event.ask_vegetarian:
            worksheet.column_dimensions[iterator.column].width = 13
            iterator.write(_("Vegetarian"), bold).right()

        if job.infection_instruction:
            worksheet.column_dimensions[iterator.column].width = 50
            iterator.write(_("Food handling"), bold).right()

        worksheet.column_dimensions[iterator.column].width = 50
        iterator.write(_("Comment"), bold).right()
        last_column = iterator.colint - 1

        iterator.column_return()
        worksheet.freeze_panes = iterator.get()

        iterator.down()

        # Add coordinators
        if not date and job.coordinators.exists():
            worksheet.merge_cells(start_row=iterator.row, start_column=1,
                                  end_row=iterator.row, end_column=last_column)

            iterator.write(_("Coordinators"), bold)
            iterator.down()
            add_helpers(worksheet, iterator, event, job,
                        job.coordinators.all(), multiple_shifts)

        # show all shifts
        for shift in job.shift_set.order_by('begin'):
            if date and shift.begin.date() != date:
                continue

            worksheet.merge_cells(start_row=iterator.row, start_column=1,
                                  end_row=iterator.row, end_column=last_column)
            iterator.write(shift.time(), bold)
            iterator.down()

            add_helpers(worksheet, iterator, event, job,
                        shift.helper_set.all(), multiple_shifts)

    workbook.save(buffer)


def add_helpers(worksheet, iterator, event, job, helpers, multiple_shifts_format):
    for helper in helpers:
        num_shifts = helper.shifts.count()
        num_jobs = len(helper.coordinated_jobs)
        fmt = None
        if num_shifts + num_jobs > 1:
            fmt = multiple_shifts_format

        iterator.write(escape(helper.firstname), fmt).right()
        iterator.write(escape(helper.surname), fmt).right()
        iterator.write(escape(helper.email), fmt).right()
        iterator.write(escape(helper.phone), fmt).right()

        if event.ask_shirt:
            iterator.write(escape(str(helper.get_shirt_display())), fmt).right()
        if event.ask_vegetarian:
            iterator.write(escape(filters.yesno(helper.vegetarian)), fmt).right()
        if job.infection_instruction:
            iterator.write(escape(str(helper.get_infection_instruction_short())), fmt).right()

        iterator.write(escape(helper.comment), fmt)
        iterator.down()
