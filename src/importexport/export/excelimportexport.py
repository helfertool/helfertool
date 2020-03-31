import re
import datetime

from django.core.exceptions import ObjectDoesNotExist
from django.utils import timezone
from django.utils.translation import ugettext as _

import openpyxl

from .xlsxutils import CellIterator, cleanName


def xlsx_generate_import_template(buffer, event, job):
    # some helpers
    def write_datetime(iterator, dt):
        # Problem: .date is in UTC, .time is in UTC.
        # We want to do it in the stored timezone
        current_tz = timezone.get_current_timezone()
        local = current_tz.normalize(dt)

        iterator.write(local.date()).right()
        iterator.write(local.time())

    # create xlsx
    workbook = openpyxl.Workbook()
    workbook.active.title = cleanName(job.name)[:20]

    # add in all the reference data into the 'data' sheet as key-value store in Colums A and B
    # this data is added for later sanity checking that, even if somebody modified these fields
    # by copying them from last year, this has to be adjusted as well, making it unreasonable to
    # actually do.
    datasheet = workbook.create_sheet(title='helfertool_internal')
    datasheet.append(['property', 'value'])
    datasheet.append(['version', "0.1"])
    datasheet.append(['event', event.url_name])
    datasheet.append(['eventname', event.name])
    datasheet.append(['eventstart', event.date])
    datasheet.append(['job', job.id])
    datasheet.append(['name', job.name])
    datasheet.append(['jobsheet', workbook.active.title])

    # avoid a user *accidentally* modifying the data structures here
    # using a simple local write lock (no real protection, just against accidents)
    datasheet.protection.enable()

    # put the structure and the template into the active first sheet
    ws = workbook.active
    iterator = CellIterator(ws)
    bold = openpyxl.styles.Font(bold=True)

    # give some documentation for the user
    iterator.write(job.name, bold).down()
    iterator.write(_("Please fill out the following form. Add dates and times seperately and do "
                     "not mess with the formatting or row/column alignment.")).down()
    iterator.write(_("Select or more with giftsets seperated with ';' "))
    iterator.write(_("Do not leave empty rows!")).down()
    iterator.write(_("Available Giftsets:")).right()
    possible_giftsets = "; ".join([giftset.name for giftset in event.giftset_set.all()])
    iterator.write(possible_giftsets).right()

    iterator.down()
    iterator.down()

    # set up header
    ws.column_dimensions[iterator.column].width = 18
    iterator.write(_("Begin Date"), bold).right()
    ws.column_dimensions[iterator.column].width = 12
    iterator.write(_("Begin Time"), bold).right()

    ws.column_dimensions[iterator.column].width = 18
    iterator.write(_("End Date"), bold).right()
    ws.column_dimensions[iterator.column].width = 12
    iterator.write(_("End Time"), bold).right()

    ws.column_dimensions[iterator.column].width = 18
    iterator.write(_("Number of helpers"), bold).right()
    ws.column_dimensions[iterator.column].width = 17
    iterator.write(_("Name (optional)"), bold).right()

    ws.column_dimensions[iterator.column].width = 16
    iterator.write(_("Hidden (Y/N)"), bold).right()

    ws.column_dimensions[iterator.column].width = 18
    iterator.write(_("Displayed as full (Y/N)"), bold).right()

    ws.column_dimensions[iterator.column].width = 17
    iterator.write(_("Giftsets"), bold).right()

    ws.column_dimensions[iterator.column].hidden = True
    iterator.write('internal', bold).right()

    iterator.down()

    # setup data validation. dunno if this actually works?
    # does not work in libreoffice.
    # does not work in gnumeric.
    date_validation = openpyxl.worksheet.datavalidation.DataValidation(type="date")
    time_validation = openpyxl.worksheet.datavalidation.DataValidation(type="time")
    integer_validation = openpyxl.worksheet.datavalidation.DataValidation(type="whole")
    yn_validation = openpyxl.worksheet.datavalidation.DataValidation(
        type="list",
        formula1='"y,Y,n,N"',
        allow_blank=True)

    date_validation.add(iterator.range_to(None, -1))  # Start date
    iterator.right()
    time_validation.add(iterator.range_to(None, -1))  # Start time
    iterator.right()
    date_validation.add(iterator.range_to(None, -1))  # End date
    iterator.right()
    time_validation.add(iterator.range_to(None, -1))  # End time
    iterator.right()
    integer_validation.add(iterator.range_to(None, -1))  # Number of helpers
    iterator.right()
    # name
    iterator.right()
    yn_validation.add(iterator.range_to(None, -1))  # hidden from public
    iterator.right()
    yn_validation.add(iterator.range_to(None, -1))  # displayed as full
    iterator.right()
    # giftsets (no I will not create a validator for that... too lazy)
    iterator.right()
    # internal
    iterator.right()

    iterator.column_return()
    datasheet.append(['data_start', iterator.get()])

    # if no shift exists, add a dummy one
    if not job.shift_set.exists():
        iterator.write(event.date).right()
        iterator.write(datetime.time.fromisoformat('12:00')).right()
        iterator.write(event.date).right()
        iterator.write(datetime.time.fromisoformat('13:00')).right()
        iterator.write('10').right()
        iterator.write('CHANGEME').right()
        iterator.write('Y').right()
        iterator.write('Y').right()
        iterator.write('None').right()
        iterator.down()

    # now add all shifts
    for shift in job.shift_set.all():
        write_datetime(iterator, shift.begin)
        iterator.right()
        write_datetime(iterator, shift.end)
        iterator.right()
        iterator.write(shift.number).right()
        iterator.write(shift.name).right()
        iterator.write('Y' if shift.hidden else 'N').right()
        iterator.write('Y' if shift.blocked else 'N').right()

        shift_giftsets = "; ".join([giftset.name for giftset in shift.gifts.all()])
        iterator.write(shift_giftsets).right()

        iterator.write("id: {}".format(shift.id)).right()

        iterator.down()

    # and finally save it
    workbook.save(buffer)


def xlsx_read_import_template(buffer, event):
    """
    Read in the document returned by the user.

    Note: import will not verify, that the giftsets exist

    :return: (errors: [<str>],
        job,
        existing_shifts:
        new_shifts: [{
            'begin': datetime,
            'end': datetime,
            'number': int,
            'name': str,
            'hidden': bool,
            'blocked': bool,
            'gifts':[str]
        }]
    )
    """
    # a few helpers
    def load_datetime(iterator):
        """ Load a datetime object from to seperate columns """
        date = iterator.cell.value
        if isinstance(date, datetime.date):
            pass
        elif isinstance(date, datetime.datetime):
            date = date.date()
        elif isinstance(date, str):
            date = datetime.date.fromisoformat(date)
        else:
            raise ValueError("Cannot handle type " + str(type(date)) + " as date")

        iterator.right()
        time = iterator.cell.value
        if isinstance(time, datetime.time):
            pass
        elif isinstance(time, datetime.datetime):
            time = time.time()
        elif isinstance(time, str):
            time = datetime.time.fromisoformat(time)
        else:
            raise ValueError("Cannot handle type " + str(type(time)) + " as time")

        return timezone.make_aware(datetime.datetime.combine(date, time))

    def load_yn(iterator):
        """ Convert Y/N to bool """
        if iterator.cell.value in ['Y', 'y', 'J', 'j', '1', 'true', 'ja', 'yes']:
            return True
        elif iterator.cell.value in ['N', 'n', '0', 'false', 'no', 'nein']:
            return False
        else:
            return False

    # Open the workbook
    errors = []
    workbook = openpyxl.load_workbook(buffer)

    # read the data tab
    datasheet = workbook['helfertool_internal']
    iterator = CellIterator(datasheet)
    fields = {}
    while iterator.cell.value:
        k = iterator.cell.value
        iterator.right()
        v = iterator.cell.value
        fields[k] = v
        iterator.down()

    # verify the fields in the data tab
    # To Check: Should we use all fields for verification making it an import error if the
    # name of the event changed in the meantime?
    try:
        if fields['version'] != "0.1":
            errors.append("Invalid export version {}".format(fields["version"]))
        if fields['event'] != event.url_name:
            errors.append("Importing to the wrong event.")

        try:

            job = event.job_set.get(event=event, id=fields['job'])
        except ObjectDoesNotExist:
            errors.append("No job with that ID was found")

        # internationalization is a bitch
        # elif job.get().name != fields['name']:
        #    errors.append("the Job you are importing has a different name")

        if fields['jobsheet'] not in workbook:
            errors.append("the jobsheet does not exist in the workbook")

        if not re.match(r"^[A-Z]+[1-9][0-9]*$", fields['data_start']):
            errors.append("data_start is invalid")

    except KeyError:
        errors.append("Could not verify data: Missing data")

    if errors:
        return errors, None, None, None

    # prepare the jobsheet
    worksheet = workbook[fields['jobsheet']]
    iterator = CellIterator(worksheet)
    iterator.load(fields['data_start'])

    # and start reading the data
    existing_shifts = []
    new_shifts = []

    while iterator.cell.value:
        shift = {}
        # Columns (in that order):
        # Start data
        # Start time
        # Stop date
        # Stop time
        # Number of helpers
        # Name
        # hidden
        # full
        # giftsets
        # internal
        try:
            shift['excel_row'] = iterator.row
            shift['begin'] = load_datetime(iterator)
            iterator.right()
            shift['end'] = load_datetime(iterator)
            iterator.right()

            shift['number'] = int(iterator.cell.value)
            iterator.right()
            shift['name'] = iterator.cell.value or ""
            iterator.right()

            shift['hidden'] = load_yn(iterator)
            iterator.right()
            shift['blocked'] = load_yn(iterator)
            iterator.right()

            gifts = (iterator.cell.value or "").split(';')
            gifts = [g.strip() for g in gifts]
            shift['giftsets'] = gifts
            iterator.right()

            internal = iterator.cell.value
            # cell is None or empty when user created a new row
            if not internal:
                shift['id'] = None
            elif 'invalid' in internal:
                continue
            elif 'id:' in internal:
                try:
                    shift['id'] = int(internal[3:])
                except ValueError:
                    errors.append(iterator.get() + ': id is not parseable')
                    continue
            iterator.right()

            if shift['id'] is not None:
                existing_shifts.append(shift)
            else:
                new_shifts.append(shift)
        except ValueError as e:
            errors.append("At " + str(iterator.get()) + ": " + str(e))
        iterator.down()
    return errors, job, existing_shifts, new_shifts
